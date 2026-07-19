import asyncio
import csv
import io
import re
import traceback

import httpx
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from pydantic import BaseModel

from app.auth import CurrentUser, require_workspace_admin
from app.config import settings
from app.db import get_supabase

router = APIRouter()

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Header aliases so admins can upload whatever their HR export happens to
# call each column, matched case-insensitively after stripping whitespace.
EMAIL_HEADERS = {"email", "email address", "e-mail", "work email"}
NAME_HEADERS = {"name", "full name", "employee name"}
ROLE_HEADERS = {"role", "access", "permission"}


# The caller's identity is never taken from the request body — it comes from
# the verified token, so these carry only the thing being acted on.
class InviteRequest(BaseModel):
    workspace_id: str
    email: str
    role: str = "member"


class RemoveMemberRequest(BaseModel):
    workspace_id: str
    target_user_id: str


class UpdateWorkspaceRequest(BaseModel):
    workspace_id: str
    name: str


class UpdateMemberRoleRequest(BaseModel):
    workspace_id: str
    target_user_id: str
    role: str


class CancelInviteRequest(BaseModel):
    workspace_id: str
    invite_id: str


class UpdateMemberDetailsRequest(BaseModel):
    workspace_id: str
    target_user_id: str
    full_name: str | None = None
    phone: str | None = None


def _is_only_admin(supabase, workspace_id: str, user_id: str) -> bool:
    # limit(1) rather than single(): the target may not be a member at all,
    # and single() raises on an empty match, which would surface as a 500.
    target = (
        supabase.table("workspace_members")
        .select("role")
        .eq("workspace_id", workspace_id)
        .eq("user_id", user_id)
        .limit(1)
        .execute()
    )
    rows = target.data or []
    if not rows or rows[0].get("role") != "admin":
        return False

    admin_count = (
        supabase.table("workspace_members")
        .select("id", count="exact")
        .eq("workspace_id", workspace_id)
        .eq("role", "admin")
        .execute()
    )
    return (admin_count.count or 0) <= 1


def _workspace_name(supabase, workspace_id: str) -> str:
    workspace = (
        supabase.table("workspaces")
        .select("name")
        .eq("id", workspace_id)
        .single()
        .execute()
    )
    return workspace.data.get("name", "your team") if workspace.data else "your team"


async def _send_invite_email(email: str, role: str, workspace_name: str) -> None:
    if not settings.resend_api_key:
        return
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            await client.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {settings.resend_api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "from": settings.resend_from_email,
                    "to": [email],
                    "subject": f"You've been invited to join {workspace_name} on Vivran.ai",
                    "html": (
                        f"<p>You've been invited to join <b>{workspace_name}</b> on Vivran.ai "
                        f"as a <b>{role}</b>.</p>"
                        f"<p>Sign in or create an account using <b>{email}</b> — "
                        f"the invite is matched on that address, so using a different "
                        f"one won't join you to the team.</p>"
                        f'<p><a href="{settings.app_url}/login" '
                        f'style="display:inline-block;padding:10px 18px;background:#4f46e5;'
                        f'color:#ffffff;border-radius:8px;text-decoration:none;'
                        f'font-weight:600">Join {workspace_name}</a></p>'
                        f'<p style="color:#71717a;font-size:12px">'
                        f'Or paste this into your browser: {settings.app_url}/login</p>'
                    ),
                },
            )
    except Exception:
        pass


async def _invite_one(
    supabase,
    workspace_id: str,
    inviter_user_id: str,
    email: str,
    role: str,
    workspace_name: str,
) -> dict:
    """Queue a single invite. Returns {"email", "status": "invited"|"skipped", "reason"?}."""
    existing_member = (
        supabase.table("workspace_members")
        .select("id")
        .eq("workspace_id", workspace_id)
        .eq("email", email)
        .execute()
    )
    if existing_member.data:
        return {"email": email, "status": "skipped", "reason": "already a member"}

    supabase.table("workspace_invites").upsert(
        {
            "workspace_id": workspace_id,
            "email": email,
            "role": role,
            "invited_by": inviter_user_id,
            "status": "pending",
        },
        on_conflict="workspace_id,email",
    ).execute()

    await _send_invite_email(email, role, workspace_name)
    return {"email": email, "status": "invited"}


@router.post("/team/invite")
async def invite_member(req: InviteRequest, user: CurrentUser):
    await require_workspace_admin(req.workspace_id, user)

    supabase = get_supabase()

    if req.role not in ("admin", "member"):
        raise HTTPException(status_code=400, detail="Invalid role")

    workspace_name = _workspace_name(supabase, req.workspace_id)

    result = await _invite_one(
        supabase, req.workspace_id, user.id, req.email, req.role, workspace_name
    )
    if result["status"] == "skipped":
        raise HTTPException(status_code=400, detail="This person is already a member")

    return {"status": "invited", "email": req.email}


@router.post("/team/accept-invites")
async def accept_invites(user: CurrentUser):
    """Turn any pending invites addressed to the caller into memberships.

    Called on sign-in rather than only at sign-up. A signup-time hook would
    miss the common case of someone who already has an account being invited
    to a workspace later — they'd get the email, sign in, and still not be a
    member of anything.

    Idempotent: an invite for a workspace the caller already belongs to is
    marked accepted without touching the membership.
    """
    if not user.email:
        return {"joined": [], "workspace_ids": []}

    email = user.email.strip().lower()
    supabase = get_supabase()

    def run() -> list[str]:
        # ilike, not eq: single invites store whatever case the admin typed
        # (only the bulk importer lowercases), so an exact match would silently
        # skip them.
        invites = (
            supabase.table("workspace_invites")
            .select("id, workspace_id, role")
            .ilike("email", email)
            .eq("status", "pending")
            .execute()
        )

        joined: list[str] = []
        for invite in invites.data or []:
            workspace_id = invite["workspace_id"]

            existing = (
                supabase.table("workspace_members")
                .select("id")
                .eq("workspace_id", workspace_id)
                .eq("user_id", user.id)
                .limit(1)
                .execute()
            )
            if not (existing.data or []):
                supabase.table("workspace_members").insert({
                    "workspace_id": workspace_id,
                    "user_id": user.id,
                    "email": email,
                    "role": invite.get("role") or "member",
                }).execute()
                joined.append(workspace_id)

            supabase.table("workspace_invites").update(
                {"status": "accepted"}
            ).eq("id", invite["id"]).execute()

        return joined

    try:
        joined = await asyncio.to_thread(run)
    except Exception:
        # Never block sign-in on this — the caller fires it on every session
        # start and ignores the result.
        print(f"[invites] Accepting invites for {email} failed:")
        traceback.print_exc()
        return {"joined": 0, "workspace_ids": []}

    if joined:
        print(f"[invites] {email} joined {len(joined)} workspace(s)")

    return {"joined": len(joined), "workspace_ids": joined}


def _parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    return _rows_to_dicts(header, rows)


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = iter(reader)
    header = next(rows, None)
    if not header:
        return []
    return _rows_to_dicts(header, rows)


def _rows_to_dicts(header, rows) -> list[dict]:
    col_index: dict[str, int] = {}
    for i, cell in enumerate(header):
        key = str(cell or "").strip().lower()
        if key in EMAIL_HEADERS:
            col_index["email"] = i
        elif key in NAME_HEADERS:
            col_index["name"] = i
        elif key in ROLE_HEADERS:
            col_index["role"] = i

    if "email" not in col_index:
        raise HTTPException(
            status_code=400,
            detail="Couldn't find an email column. Expected a header like 'Email'.",
        )

    parsed = []
    for row in rows:
        if row is None or all(cell in (None, "") for cell in row):
            continue

        def cell_at(key: str) -> str:
            idx = col_index.get(key)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        parsed.append(
            {
                "email": cell_at("email"),
                "name": cell_at("name"),
                "role": cell_at("role"),
            }
        )
    return parsed


@router.post("/team/bulk-invite")
async def bulk_invite_members(
    user: CurrentUser,
    workspace_id: str = Form(...),
    file: UploadFile = File(...),
):
    await require_workspace_admin(workspace_id, user)

    supabase = get_supabase()

    filename = (file.filename or "").lower()
    content = await file.read()

    if filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail="Upload a .xlsx or .csv file")

    if not rows:
        raise HTTPException(status_code=400, detail="No employee rows found in the file")

    workspace_name = _workspace_name(supabase, workspace_id)

    invited: list[str] = []
    skipped: list[dict] = []
    seen_emails: set[str] = set()

    for row in rows:
        email = row["email"].lower()
        if not email or not EMAIL_RE.match(email):
            skipped.append({"email": row["email"] or "(blank)", "reason": "invalid email"})
            continue
        if email in seen_emails:
            skipped.append({"email": email, "reason": "duplicate row"})
            continue
        seen_emails.add(email)

        role = row["role"].lower()
        if role not in ("admin", "member"):
            role = "member"

        result = await _invite_one(supabase, workspace_id, user.id, email, role, workspace_name)
        if result["status"] == "invited":
            invited.append(email)
        else:
            skipped.append({"email": email, "reason": result["reason"]})

    return {"invited": invited, "skipped": skipped, "total_rows": len(rows)}


@router.post("/team/remove-member")
async def remove_member(req: RemoveMemberRequest, user: CurrentUser):
    await require_workspace_admin(req.workspace_id, user)

    supabase = get_supabase()

    if _is_only_admin(supabase, req.workspace_id, req.target_user_id):
        raise HTTPException(status_code=400, detail="Cannot remove the only admin")

    supabase.table("workspace_members").delete().eq("workspace_id", req.workspace_id).eq(
        "user_id", req.target_user_id
    ).execute()

    return {"status": "removed"}


@router.post("/team/update-workspace")
async def update_workspace(req: UpdateWorkspaceRequest, user: CurrentUser):
    await require_workspace_admin(req.workspace_id, user)

    supabase = get_supabase()

    name = req.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Workspace name can't be empty")
    if len(name) > 80:
        raise HTTPException(status_code=400, detail="Workspace name is too long")

    supabase.table("workspaces").update({"name": name}).eq("id", req.workspace_id).execute()

    return {"status": "ok", "name": name}


@router.post("/team/update-member-role")
async def update_member_role(req: UpdateMemberRoleRequest, user: CurrentUser):
    await require_workspace_admin(req.workspace_id, user)

    supabase = get_supabase()

    if req.role not in ("admin", "member"):
        raise HTTPException(status_code=400, detail="Invalid role")

    if req.role == "member" and _is_only_admin(supabase, req.workspace_id, req.target_user_id):
        raise HTTPException(status_code=400, detail="Cannot demote the only admin")

    supabase.table("workspace_members").update({"role": req.role}).eq(
        "workspace_id", req.workspace_id
    ).eq("user_id", req.target_user_id).execute()

    return {"status": "ok", "role": req.role}


@router.post("/team/update-member-details")
async def update_member_details(req: UpdateMemberDetailsRequest, user: CurrentUser):
    await require_workspace_admin(req.workspace_id, user)

    supabase = get_supabase()

    membership = (
        supabase.table("workspace_members")
        .select("id")
        .eq("workspace_id", req.workspace_id)
        .eq("user_id", req.target_user_id)
        .execute()
    )
    if not membership.data:
        raise HTTPException(status_code=404, detail="Member not found in this workspace")

    full_name = req.full_name.strip() if req.full_name is not None else None
    if full_name is not None:
        if not full_name:
            raise HTTPException(status_code=400, detail="Name can't be empty")
        supabase.table("profiles").update({"full_name": full_name}).eq(
            "id", req.target_user_id
        ).execute()

    if req.phone is not None:
        phone = req.phone.strip() or None
        supabase.table("workspace_members").update({"phone": phone}).eq(
            "workspace_id", req.workspace_id
        ).eq("user_id", req.target_user_id).execute()

    return {"status": "ok", "full_name": full_name, "phone": req.phone}


@router.post("/team/cancel-invite")
async def cancel_invite(req: CancelInviteRequest, user: CurrentUser):
    await require_workspace_admin(req.workspace_id, user)

    supabase = get_supabase()

    supabase.table("workspace_invites").delete().eq("id", req.invite_id).eq(
        "workspace_id", req.workspace_id
    ).execute()

    return {"status": "cancelled"}
